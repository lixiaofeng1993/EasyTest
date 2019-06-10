# coding:utf-8
from openpyxl import load_workbook
import openpyxl
from openpyxl.styles import Font, colors


def copy_excel(cese_path, report_path):
    """复制测试用例到report_path"""
    wb2 = openpyxl.Workbook()
    wb2.save(report_path)  # 在设置的路径下创建一个excel文件
    # 读取数据
    wb1 = openpyxl.load_workbook(cese_path)
    wb2 = openpyxl.load_workbook(report_path)
    sheets1 = wb1.sheetnames
    sheets2 = wb2.sheetnames
    sheet1 = wb1[sheets1[0]]  # 获取第一个sheet页
    sheet2 = wb2[sheets2[0]]
    max_row = sheet1.max_row  # 最大行数
    max_column = sheet1.max_column  # 最大列数

    for m in list(range(1, max_row + 1)):
        for n in list(range(97, 97 + max_column)):  # chr(97)='a'
            n = chr(n)  # ASCII字符,excel文件的列 a b c
            i = '%s%d' % (n, m)  # 单元格编号
            cell1 = sheet1[i].value  # 获取测试用例单元格数据
            sheet2[i].value = cell1  # 赋值到测试结果单元格

    wb2.save(report_path)  # 保存数据
    wb1.close()  # 关闭excel
    wb2.close()


class Write_excel(object):
    """修改excel数据"""

    def __init__(self, filename):
        self.filename = filename
        self.wb = load_workbook(self.filename)
        self.ws = self.wb.active  # 激活sheet

    def write(self, row_n, col_n, value):
        """写入数据，如(2,3，"hello"),第二行第三列写入数据"hello\""""
        ft = Font(color=colors.RED, size=12, bold=True)
        # 判断值为错误时添加字体样式
        if value in ['fail', 'error'] or col_n == 12:
            self.ws.cell(row_n, col_n).font = ft
        if value == 'pass':
            ft = Font(color=colors.GREEN)
            self.ws.cell(row_n, col_n).font = ft
        self.ws.cell(row_n, col_n).value = value
        self.wb.save(self.filename)


if __name__ == "__main__":
    # copy_excel("demo_api_3.xlsx", "test111.xlsx")
    wt = Write_excel("test111.xlsx")
    wt.write(4, 5, "HELLEOP")
    wt.write(4, 6, "HELLEOP")
